from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timedelta
from typing import Any, Optional

from sqlalchemy import select, case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.agent import Agent
from app.models.campaign import Campaign, Project
from app.models.contact import Contact
from app.models.lead import Lead
from app.models.models import Activity
from app.services.lead_service import create_auto_task, create_notification


# All 16 fields from the Excel schema
EXPECTED_FIELDS = [
    "call_id",
    "name",
    "phone_number",
    "other_info",
    "attempt_number",
    "transcript",
    "recording_url",
    "extracted_entities",
    "call_eval_tag",
    "summary",
    "call_conversation_quality",
    "call_dialing_at",
    "call_ringing_at",
    "user_picked_up",
    "num_of_retries",
    "dial_status_reason",
]

STAGE_RANK = {
    "new": 0,
    "contacted": 1,
    "site_visit_scheduled": 2,
    "site_visit_done": 3,
    "negotiation": 4,
    "won": 5,
    "lost": 6,
    "nurture": 1,
}


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _safe_int(v: Any, default: int = 0) -> int:
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def parse_campaign_file(file_content: bytes, filename: str) -> tuple[list[dict], str]:
    """Parse CSV, JSON, or XLSX campaign files into normalised row dicts."""
    ext = (filename.rsplit(".", 1)[-1].lower() if "." in filename else "")
    rows: list[dict] = []

    if ext == "csv":
        text = file_content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            lowered = {(_safe_str(k).lower()): v for k, v in (row or {}).items()}
            rows.append({field: _safe_str(lowered.get(field, "")) for field in EXPECTED_FIELDS})
        return rows, "csv"

    if ext == "json":
        payload = json.loads(file_content.decode("utf-8", errors="replace") or "[]")
        if not isinstance(payload, list):
            raise ValueError("JSON payload must be an array of objects")
        for item in payload:
            if not isinstance(item, dict):
                continue
            rows.append({field: _safe_str(item.get(field, "")) for field in EXPECTED_FIELDS})
        return rows, "json"

    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl is required for Excel file support. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)

        # Choose the most suitable worksheet for ingestion.
        # Preference order:
        # 1) Sheet names containing "raw" with expected headers
        # 2) Any sheet with expected headers and most rows
        # 3) First sheet fallback
        expected = set(EXPECTED_FIELDS)
        sheet_name: Optional[str] = None
        best_score: tuple[int, int, int] = (-1, -1, -1)

        for name in wb.sheetnames:
            ws_candidate = wb[name]
            try:
                header_row = next(ws_candidate.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue

            headers = {_safe_str(v).lower() for v in header_row if _safe_str(v)}
            match_count = len(headers & expected)
            row_count = max((ws_candidate.max_row or 1) - 1, 0)

            # Only treat sheet as structured campaign data if enough expected columns exist.
            has_expected_schema = match_count >= 4
            if not has_expected_schema:
                continue

            priority = 2 if "raw" in name.lower() else 1
            score = (priority, match_count, row_count)
            if score > best_score:
                best_score = score
                sheet_name = name

        if sheet_name is None:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Read headers from first row
        headers_raw = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers_raw.append(_safe_str(cell.value).lower() if cell.value else "")

        for row_cells in ws.iter_rows(min_row=2):
            row_dict: dict[str, str] = {}
            for col_idx, cell in enumerate(row_cells):
                if col_idx < len(headers_raw):
                    header = headers_raw[col_idx]
                    row_dict[header] = _safe_str(cell.value)

            # Only add non-empty rows
            if row_dict.get("name") or row_dict.get("phone_number"):
                rows.append({field: row_dict.get(field, "") for field in EXPECTED_FIELDS})

        wb.close()
        return rows, "xlsx"

    raise ValueError("Unsupported file type. Please upload .csv, .json, or .xlsx")


def normalise_phone(phone: str) -> str:
    raw = _safe_str(phone)
    digits = re.sub(r"\D", "", raw)
    if raw.startswith("+"):
        return "+" + digits
    if len(digits) == 10:
        return "+91" + digits
    if digits.startswith("91") and len(digits) == 12:
        return "+" + digits
    return "+" + digits if digits else ""


def _contains_any(text: str, terms: list[str]) -> bool:
    return any(term in text for term in terms)


def _load_entities(extracted_entities: str) -> dict:
    value = _safe_str(extracted_entities)
    if not value:
        return {}
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _load_quality(quality_str: str) -> dict:
    """Parse call_conversation_quality JSON string."""
    value = _safe_str(quality_str)
    if not value:
        return {}
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _pattern_count(text: str, patterns: list[str]) -> int:
    count = 0
    for pattern in patterns:
        if re.search(pattern, text):
            count += 1
    return count


def _signal_features(text: str) -> dict:
    hot_patterns = [
        r"\bsite\s*visit\s*(confirmed|scheduled|arranged)?\b",
        r"\bpick[- ]?up\s*(arranged|confirmed)?\b",
        r"\bfamily\s+(is\s+)?(coming|visiting|visit)\b",
        r"\b(visit|come)\s+(tomorrow|today|this\s+week)\b",
        r"\bwhatsapp\s+(details|brochure|price|pricing)\b",
        r"\b(2|3|4)\s*bhk\b",
        r"\bcarpet\s*area\b",
        r"\bloading\s*percentage\b",
        r"\bdown\s*payment\b",
        r"\bpayment\s*plan\b",
        r"\bhome\s*loan\b",
        r"\b(booked|book|finali[sz]e|confirm(ed)?)\b",
    ]

    warm_patterns = [
        r"\bcould\s*visit\b",
        r"\bmight\s*visit\b",
        r"\bwill\s*think\b",
        r"\bcheck\s*with\s*family\b",
        r"\bcall\s*back\b",
        r"\bneeds?\s*more\s*time\b",
        r"\bif\s*price\s*is\s*right\b",
        r"\binterested\b",
        r"\bcredai\s*expo\b",
    ]

    cold_patterns = [
        r"\bnot\s*interested\b",
        r"\bno\s*interest\b",
        r"\bdon['']?t\s*want\b",
        r"\balready\s*(bought|invested)\b",
        r"\bbudget\s*too\s*low\b",
        r"\bvoicemail\b",
        r"\bdid\s*not\s*answer\b",
        r"\bno\s*next\s*step\b",
        r"\bremove\b",
        r"\bdo\s*not\s*call\b",
    ]

    concrete_hot_patterns = [
        r"\bsite\s*visit\s*(confirmed|scheduled|arranged)\b",
        r"\bpick[- ]?up\s*(arranged|confirmed)\b",
        r"\b(visit|come)\s*(tomorrow|today)\b",
    ]

    return {
        "hot_count": _pattern_count(text, hot_patterns),
        "warm_count": _pattern_count(text, warm_patterns),
        "cold_count": _pattern_count(text, cold_patterns),
        "concrete_hot": _pattern_count(text, concrete_hot_patterns) > 0,
    }


def _is_connected(transcript: str) -> bool:
    """Check if a call was actually connected (has a real transcript)."""
    t = _safe_str(transcript)
    return bool(t) and t.lower() not in ("", "transcript not found", "none") and len(t) > 50


# ─── PRIORITY SCORING ENGINE (P1–P7) ────────────────────────────────────────

def compute_priority_score(
    summary: str,
    transcript: str,
    entities: dict,
    quality: dict,
    call_eval_tag: str,
    attempt_number: int,
    num_of_retries: int,
    transcript_length: int,
) -> int:
    """Compute a priority score from 0–100 based on all available signals."""
    score = 0

    if not _is_connected(transcript):
        return -1  # No-connect → P7

    # Entity-based signals
    if _safe_str(entities.get("Site_Visit_Agreed", "")).lower() == "yes":
        score += 30
    if _safe_str(entities.get("whatsapp_followup", "")).lower() == "yes":
        score += 15
    if entities.get("Configuration_Preference"):
        score += 12
    if entities.get("Budget_Estimate"):
        score += 10
    if _safe_str(entities.get("call_back_requested", "")).lower() == "yes":
        score += 8
    if _safe_str(entities.get("Senior Escalation", "")).lower() == "yes":
        score += 5

    # Quality-based signals
    overall = quality.get("overall_quality", 0)
    try:
        overall = float(overall)
    except (ValueError, TypeError):
        overall = 0.0
    if overall >= 7:
        score += 15
    elif overall >= 5:
        score += 8
    elif overall >= 3:
        score += 3

    # Engagement signals (transcript length)
    if transcript_length > 1500:
        score += 10
    elif transcript_length > 700:
        score += 6
    elif transcript_length > 300:
        score += 3

    # Eval tag
    eval_l = _safe_str(call_eval_tag).lower()
    if eval_l == "yes":
        score += 15
    elif eval_l == "no" and score == 0:
        score = max(score - 5, 0)

    # Negative signals from summary/transcript
    summary_l = _safe_str(summary).lower()
    if _contains_any(summary_l, ["not interested", "no interest"]):
        score = max(score - 10, 0)
    if _contains_any(summary_l, ["already bought", "already invested"]):
        score = max(score - 15, 0)
    if _contains_any(summary_l, ["don't call", "do not call", "remove"]):
        score = max(score - 20, 0)

    # Attempt penalty for low-scoring leads on high attempts
    if attempt_number >= 3 and score < 20:
        score = max(score - 5, 0)

    return min(score, 100)


def compute_priority_tier(score: int) -> str:
    """Map a priority score to a tier (P1–P7)."""
    if score < 0:
        return "P7"
    if score >= 60:
        return "P1"
    if score >= 40:
        return "P2"
    if score >= 25:
        return "P3"
    if score >= 10:
        return "P4"
    if score >= 1:
        return "P5"
    return "P6"


def tier_to_lead_score(tier: str) -> str:
    """Map P-tier to CRM lead score (hot/warm/cold)."""
    if tier in ("P1", "P2"):
        return "hot"
    if tier in ("P3", "P4"):
        return "warm"
    return "cold"


# ─── CLASSIFICATION (enhanced with priority scoring) ────────────────────────

def classify_lead(
    summary: str,
    transcript: str,
    call_eval_tag: str,
    extracted_entities: str,
    quality_str: str = "",
    attempt_number: int = 1,
    num_of_retries: int = 0,
) -> dict:
    """Classify a lead using both rule-based scoring and priority system.
    Returns score, stage, priority, task info, AND priority_tier + priority_score.
    """
    summary_l = _safe_str(summary).lower()
    transcript_l = _safe_str(transcript).lower()
    eval_l = _safe_str(call_eval_tag).lower()
    entities = _load_entities(extracted_entities)
    quality = _load_quality(quality_str)

    # Compute P-tier priority score
    transcript_length = len(_safe_str(transcript))
    p_score = compute_priority_score(
        summary, transcript, entities, quality,
        call_eval_tag, attempt_number, num_of_retries, transcript_length
    )
    p_tier = compute_priority_tier(p_score)

    # Original rule-based classification for CRM hot/warm/cold
    summary_features = _signal_features(summary_l)
    transcript_features = _signal_features(transcript_l)

    hot_score = (summary_features["hot_count"] * 3) + (transcript_features["hot_count"] * 2)
    warm_score = (summary_features["warm_count"] * 3) + (transcript_features["warm_count"] * 2)
    cold_score = (summary_features["cold_count"] * 3) + (transcript_features["cold_count"] * 2)

    if eval_l == "yes":
        if hot_score > 0:
            hot_score += 2
        else:
            warm_score += 2
    elif eval_l == "no":
        cold_score += 2

    if entities:
        budget = _safe_str(entities.get("budget", "")).lower()
        timeline = _safe_str(entities.get("timeline", "")).lower()
        visit = _safe_str(entities.get("site_visit", entities.get("Site_Visit_Agreed", ""))).lower()
        if budget and timeline and _contains_any(timeline, ["immediate", "1 month", "this week"]):
            hot_score += 2
        if _contains_any(visit, ["yes", "scheduled", "confirmed"]):
            hot_score += 3

    has_hot = hot_score > 0
    has_cold = cold_score > 0
    concrete_hot = summary_features["concrete_hot"] or transcript_features["concrete_hot"]

    if has_hot and has_cold and not concrete_hot:
        score = "warm"
    elif hot_score >= max(warm_score, cold_score) and hot_score > 0:
        score = "hot"
    elif cold_score > max(hot_score, warm_score):
        score = "cold"
    else:
        score = "warm"

    # Override with P-tier mapping if it produces a stronger result
    tier_score = tier_to_lead_score(p_tier)
    score_rank = {"hot": 2, "warm": 1, "cold": 0}
    if score_rank.get(tier_score, 0) > score_rank.get(score, 0):
        score = tier_score

    # Determine stage and task based on classification
    blob = f"{summary_l}\n{transcript_l}"
    if score == "hot" and (
        concrete_hot
        or _contains_any(blob, ["site visit", "visit", "pick-up", "pickup", "tomorrow", "arranged"])
    ):
        return {
            "score": "hot", "stage": "site_visit_scheduled", "priority": "high",
            "task_title": "URGENT: Confirm tomorrow's site visit — reconfirm time and pick-up details",
            "task_due_hours": 1,
            "priority_tier": p_tier, "priority_score": max(p_score, 0),
        }

    if score == "hot":
        return {
            "score": "hot", "stage": "negotiation", "priority": "high",
            "task_title": "Hot lead — follow up within 1 hour",
            "task_due_hours": 1,
            "priority_tier": p_tier, "priority_score": max(p_score, 0),
        }

    if score == "warm":
        return {
            "score": "warm", "stage": "contacted", "priority": "normal",
            "task_title": "Follow-up call — invite for site visit",
            "task_due_hours": 24,
            "priority_tier": p_tier, "priority_score": max(p_score, 0),
        }

    return {
        "score": "cold", "stage": "contacted", "priority": "low",
        "task_title": "Re-engage in 30 days",
        "task_due_hours": 24 * 30,
        "priority_tier": p_tier, "priority_score": max(p_score, 0),
    }


# ─── DEDUP & PROJECT LINKING ────────────────────────────────────────────────

def _levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)

    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            ins = curr[j - 1] + 1
            delete = prev[j] + 1
            sub = prev[j - 1] + (0 if ca == cb else 1)
            curr.append(min(ins, delete, sub))
        prev = curr
    return prev[-1]


async def find_existing_lead(phone: str, name: str, db: AsyncSession) -> Optional[Lead]:
    if phone:
        result = await db.execute(
            select(Lead)
            .join(Contact, Contact.id == Lead.contact_id)
            .options(selectinload(Lead.contact))
            .where(Contact.phone == phone)
            .order_by(Lead.updated_at.desc())
        )
        lead = result.scalars().first()
        if lead:
            return lead

    name_n = _safe_str(name).lower()
    if not name_n:
        return None

    result = await db.execute(
        select(Lead)
        .join(Contact, Contact.id == Lead.contact_id)
        .options(selectinload(Lead.contact))
        .order_by(Lead.updated_at.desc())
        .limit(200)
    )
    for lead in result.scalars().all():
        contact_name = _safe_str(getattr(lead.contact, "name", "")).lower()
        if contact_name and _levenshtein(contact_name, name_n) <= 2:
            return lead
    return None


async def auto_link_project(campaign_name: str, db: AsyncSession) -> Optional[str]:
    value = _safe_str(campaign_name).lower()
    if not value:
        return None

    result = await db.execute(select(Project))
    for project in result.scalars().all():
        pname = _safe_str(project.name).lower()
        if pname and pname in value:
            return project.id
    return None


def _stage_forward_only(current_stage: str, new_stage: str) -> str:
    current_rank = STAGE_RANK.get(current_stage, 0)
    new_rank = STAGE_RANK.get(new_stage, 0)
    return new_stage if new_rank >= current_rank else current_stage


async def _resolve_notification_agent_id(lead: Lead, db: AsyncSession) -> Optional[str]:
    if lead.assigned_to:
        return lead.assigned_to
    result = await db.execute(
        select(Agent.id)
        .where(Agent.is_active == True)
        .where(Agent.role.in_(["admin", "manager"]))
        .order_by(Agent.created_at.asc())
        .limit(1)
    )
    row = result.first()
    return row[0] if row else None


# ─── PROCESS CAMPAIGN ROW (extended with all 16 fields) ─────────────────────

async def process_campaign_row(row: dict, campaign: Campaign, db: AsyncSession) -> dict:
    name = _safe_str(row.get("name"))
    phone = normalise_phone(row.get("phone_number", ""))
    summary = _safe_str(row.get("summary"))
    transcript = _safe_str(row.get("transcript"))
    recording_url = _safe_str(row.get("recording_url"))
    call_eval_tag = _safe_str(row.get("call_eval_tag"))
    extracted_entities = _safe_str(row.get("extracted_entities"))
    quality_str = _safe_str(row.get("call_conversation_quality"))
    attempt_number = _safe_int(row.get("attempt_number"), 1)
    num_of_retries = _safe_int(row.get("num_of_retries"), 0)
    call_dialing_at = _safe_str(row.get("call_dialing_at")) or None
    call_ringing_at = _safe_str(row.get("call_ringing_at")) or None
    user_picked_up = _safe_str(row.get("user_picked_up")) or None
    dial_status_reason = _safe_str(row.get("dial_status_reason"))
    other_info = _safe_str(row.get("other_info"))

    classification = classify_lead(
        summary, transcript, call_eval_tag, extracted_entities,
        quality_str, attempt_number, num_of_retries
    )

    existing = await find_existing_lead(phone, name, db)
    action = "updated" if existing else "created"

    if existing:
        lead = existing
        contact = await db.get(Contact, lead.contact_id)
        if contact:
            if name and (not contact.name or contact.name.lower() == "unknown"):
                contact.name = name
            if phone and contact.phone != phone:
                contact.phone = phone

        lead.lead_score = classification["score"]
        lead.priority = classification["priority"]
        lead.stage = _stage_forward_only(lead.stage, classification["stage"])
        lead.updated_at = datetime.utcnow()
        lead.campaign_id = campaign.id
        current_project_ids = list(lead.project_ids or [])
        if campaign.project_id and campaign.project_id not in current_project_ids:
            current_project_ids.append(campaign.project_id)
            lead.project_ids = current_project_ids

    else:
        contact = Contact(
            name=name or "Unknown",
            phone=phone,
            email=None,
            type="buyer",
            source=f"Campaign — {campaign.name}",
        )
        db.add(contact)
        await db.flush()

        lead = Lead(
            contact_id=contact.id,
            source="campaign",
            stage=classification["stage"],
            lead_score=classification["score"],
            assigned_to=None,
            priority=classification["priority"],
            campaign_id=campaign.id,
            project_ids=[campaign.project_id] if campaign.project_id else [],
            stage_changed_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(lead)
        await db.flush()

    # Build rich meta with all extended fields
    entities_parsed = _load_entities(extracted_entities)
    quality_parsed = _load_quality(quality_str)

    activity_meta = {
        "call_id": _safe_str(row.get("call_id")),
        "campaign_name": campaign.name,
        "extracted_entities": entities_parsed,
        "call_conversation_quality": quality_parsed,
        "attempt_number": attempt_number,
        "num_of_retries": num_of_retries,
        "call_dialing_at": call_dialing_at,
        "call_ringing_at": call_ringing_at,
        "user_picked_up": user_picked_up,
        "dial_status_reason": dial_status_reason,
        "other_info": other_info,
        "priority_tier": classification["priority_tier"],
        "priority_score": classification["priority_score"],
        "transcript_length": len(transcript),
        "is_connected": _is_connected(transcript),
    }

    activity = Activity(
        lead_id=lead.id,
        contact_id=lead.contact_id,
        type="campaign_call",
        title=f"Campaign call ingested: {campaign.name}",
        description=summary[:500] if summary else None,
        performed_by=None,
        performed_at=datetime.utcnow(),
        campaign_id=campaign.id,
        recording_url=recording_url or None,
        transcript=transcript or None,
        call_summary=summary or None,
        call_eval_tag=call_eval_tag or None,
        meta=activity_meta,
    )
    db.add(activity)

    await create_auto_task(
        db,
        lead_id=lead.id,
        title=classification["task_title"],
        task_type="call",
        assigned_to=lead.assigned_to,
        hours_from_now=int(classification["task_due_hours"]),
        priority=classification["priority"],
    )

    if classification["score"] == "hot":
        notify_agent_id = await _resolve_notification_agent_id(lead, db)
        if notify_agent_id:
            display_name = name or (contact.name if contact else "Lead")
            await create_notification(
                db,
                agent_id=notify_agent_id,
                title=f"HOT LEAD: {display_name} — {campaign.name} — action required",
                body=f"Campaign call classified as HOT (Tier {classification['priority_tier']}, Score {classification['priority_score']}). Immediate action recommended.",
                notif_type="new_lead",
                link=f"/leads/{lead.id}",
            )

    await db.flush()

    return {
        "action": action,
        "lead_id": lead.id,
        "score": classification["score"],
        "stage": lead.stage,
        "priority": lead.priority,
        "priority_tier": classification["priority_tier"],
        "priority_score": classification["priority_score"],
        "name": name or (contact.name if contact else "Unknown"),
        "phone": phone,
        "summary": summary,
    }


async def list_campaign_leads(campaign_id: str, db: AsyncSession) -> list[Lead]:
    result = await db.execute(
        select(Lead)
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .where(Lead.campaign_id == campaign_id)
        .order_by(
            case(
                (Lead.lead_score == "hot", 0),
                (Lead.lead_score == "warm", 1),
                else_=2,
            ),
            Lead.updated_at.desc(),
        )
    )
    return result.scalars().all()
