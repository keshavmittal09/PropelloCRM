from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Header, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, asc, desc
from sqlalchemy.orm import selectinload
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from pydantic import BaseModel
from app.core.dependencies import get_db, get_current_user
from app.core.config import settings
from app.models.agent import Agent
from app.models.contact import Contact
from app.models.lead import Lead
from app.models.models import Activity, Task
from pydantic import Field
from app.schemas.schemas import (
    InboundLead, InboundLeadResponse, LeadCreate, LeadUpdate,
    LeadResponse, StageUpdate, NoteCreate, CallLogCreate, ActivityResponse,
    MasterProfileUpdate, DemographicsInput, DemographicsResponse,
)
from app.services.services import is_sales_scoped_admin, live_sales_agent_ids
from app.services.lead_service import process_inbound_lead, change_lead_stage, log_activity, create_auto_task, create_notification
from app.services.services import find_matching_properties, send_whatsapp
from app.services.memory_service import build_memory_brief
from app.services.email_service import send_email
from app.schemas.schemas import WhatsAppSend, PropertyResponse, LeadNotifyRequest

router = APIRouter()


def _normalize_to_utc_naive_from_local(dt_value: datetime) -> datetime:
    tz = ZoneInfo(settings.NOTIFICATION_TIMEZONE)
    aware = dt_value.astimezone(tz) if dt_value.tzinfo else dt_value.replace(tzinfo=tz)
    return aware.astimezone(timezone.utc).replace(tzinfo=None)


class LeadPageResponse(BaseModel):
    items: list[LeadResponse]
    total: int
    page: int
    page_size: int
    total_pages: int


def _ensure_lead_scope(current_user: Agent, lead: Lead) -> None:
    """
    Agents and call_agents can only access leads assigned to them.
    Admin/manager can access all leads.
    """
    if current_user.role in {"agent", "call_agent"} and lead.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to access this lead")


@router.post("/inbound", response_model=InboundLeadResponse)
async def inbound_lead(
    data: InboundLead,
    db: AsyncSession = Depends(get_db),
    x_priya_secret: Optional[str] = Header(None),
):
    """
    PUBLIC endpoint — receives leads from all external sources.
    Priya AI, website form, n8n workflows all POST here.
    """
    # Optional secret header verification for Priya
    if x_priya_secret and x_priya_secret != settings.PRIYA_WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Invalid webhook secret")

    result = await process_inbound_lead(db, data)
    return InboundLeadResponse(**result)


def _apply_lead_filters(query, filters_list):
    """Apply a list of filter expressions to a SQLAlchemy query."""
    if filters_list:
        query = query.where(*filters_list)
    return query


def _build_lead_filters(
    stage: Optional[str],
    source: Optional[str],
    lead_score: Optional[str],
    assigned_to: Optional[str],
    campaign_id: Optional[str],
    sentiment: Optional[str],
    whatsapp_status: Optional[str],
    assigned: Optional[str],
    retry: Optional[str],
    min_score: Optional[int],
    max_score: Optional[int],
    date_filter: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    current_user_role: str,
    current_user_id: str,
):
    """Build the full list of SQLAlchemy filter expressions from query params."""
    filters = []

    if stage:
        filters.append(Lead.stage == stage)
    if source:
        filters.append(Lead.source == source)
    if lead_score:
        filters.append(Lead.lead_score == lead_score)
    if assigned_to:
        filters.append(Lead.assigned_to == assigned_to)
    if campaign_id:
        filters.append(Lead.campaign_id == campaign_id)

    # New filters
    if sentiment:
        filters.append(Lead.call_sentiment == sentiment)
    if whatsapp_status:
        filters.append(Lead.whatsapp_status == whatsapp_status)
    if assigned == "assigned":
        filters.append(Lead.assigned_to.isnot(None))
    elif assigned == "unassigned":
        filters.append(Lead.assigned_to.is_(None))
    if retry == "retry_1":
        filters.append(Lead.retry_count == 1)
    elif retry == "retry_2":
        filters.append(Lead.retry_count == 2)
    elif retry == "retry_3":
        filters.append(Lead.retry_count == 3)
    elif retry == "max_reached":
        filters.append(Lead.max_retries_reached == True)
    if min_score is not None:
        filters.append(Lead.call_score >= min_score)
    if max_score is not None:
        filters.append(Lead.call_score <= max_score)

    # Date filters
    now = datetime.utcnow()
    if date_filter == "today":
        day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        filters.append(Lead.created_at >= day_start)
    elif date_filter == "yesterday":
        day_start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = now.replace(hour=0, minute=0, second=0, microsecond=0)
        filters.append(Lead.created_at >= day_start, Lead.created_at < day_end)
    elif date_filter == "this_week":
        week_start = now - timedelta(days=now.weekday())
        filters.append(Lead.created_at >= week_start.replace(hour=0, minute=0, second=0, microsecond=0))
    elif date_filter == "this_month":
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        filters.append(Lead.created_at >= month_start)
    elif date_filter == "custom":
        if date_from:
            try:
                filters.append(Lead.created_at >= datetime.fromisoformat(date_from))
            except ValueError:
                pass
        if date_to:
            try:
                filters.append(Lead.created_at <= datetime.fromisoformat(date_to))
            except ValueError:
                pass

    # Role-based scoping
    if current_user_role in ["agent", "call_agent"]:
        filters.append(Lead.assigned_to == current_user_id)

    return filters


@router.get("", response_model=list[LeadResponse])
async def list_leads(
    stage: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    lead_score: Optional[str] = Query(None),
    assigned_to: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    sentiment: Optional[str] = Query(None),
    whatsapp_status: Optional[str] = Query(None),
    assigned: Optional[str] = Query(None),
    retry: Optional[str] = Query(None),
    min_score: Optional[int] = Query(None, ge=0, le=100),
    max_score: Optional[int] = Query(None, ge=0, le=100),
    date_filter: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0),
    limit: int = Query(50),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    filters = _build_lead_filters(
        stage, source, lead_score, assigned_to, campaign_id,
        sentiment, whatsapp_status, assigned, retry,
        min_score, max_score, date_filter, date_from, date_to,
        current_user.role, current_user.id,
    )
    # Sales-scoped admins (e.g. Krishna group) only ever see the live sales
    # team's leads, never the full database.
    if is_sales_scoped_admin(current_user):
        filters = list(filters) + [Lead.assigned_to.in_(live_sales_agent_ids())]

    query = (
        select(Lead)
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .order_by(Lead.updated_at.desc())
    )

    if search:
        query = query.join(Contact).where(
            or_(Contact.name.ilike(f"%{search}%"), Contact.phone.ilike(f"%{search}%"))
        )
        if filters:
            query = query.where(*filters)
    else:
        if filters:
            query = query.where(*filters)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return [LeadResponse.model_validate(l) for l in result.scalars().all()]


@router.get("/paginated", response_model=LeadPageResponse)
async def list_leads_paginated(
    stage: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    lead_score: Optional[str] = Query(None),
    assigned_to: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    sentiment: Optional[str] = Query(None),
    whatsapp_status: Optional[str] = Query(None),
    assigned: Optional[str] = Query(None),
    retry: Optional[str] = Query(None),
    min_score: Optional[int] = Query(None, ge=0, le=100),
    max_score: Optional[int] = Query(None, ge=0, le=100),
    date_filter: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    filters = _build_lead_filters(
        stage, source, lead_score, assigned_to, campaign_id,
        sentiment, whatsapp_status, assigned, retry,
        min_score, max_score, date_filter, date_from, date_to,
        current_user.role, current_user.id,
    )
    # Sales-scoped admins (e.g. Krishna group) only ever see the live sales
    # team's leads, never the full database.
    if is_sales_scoped_admin(current_user):
        filters = list(filters) + [Lead.assigned_to.in_(live_sales_agent_ids())]

    base_query = select(Lead)
    count_query = select(func.count(Lead.id))

    if search:
        search_expr = or_(Contact.name.ilike(f"%{search}%"), Contact.phone.ilike(f"%{search}%"))
        base_query = base_query.join(Contact).where(search_expr)
        count_query = count_query.select_from(Lead).join(Contact).where(search_expr)

    if filters:
        base_query = base_query.where(*filters)
        count_query = count_query.where(*filters)

    total = (await db.execute(count_query)).scalar() or 0

    data_query = (
        base_query
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .order_by(Lead.updated_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )

    result = await db.execute(data_query)
    items = [LeadResponse.model_validate(l) for l in result.scalars().all()]

    return LeadPageResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size if total else 1,
    )


@router.get("/board", response_model=dict)
async def kanban_board(
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Returns leads grouped by stage for the Kanban board."""
    stages = ["new", "contacted", "site_visit_scheduled", "site_visit_done", "negotiation", "won", "lost", "nurture"]
    query = (
        select(Lead)
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .where(Lead.stage.in_(stages))
    )
    if current_user.role in ["agent", "call_agent"]:
        query = query.where(Lead.assigned_to == current_user.id)

    result = await db.execute(query)
    leads = result.scalars().all()

    board = {stage: [] for stage in stages}
    for lead in leads:
        board[lead.stage].append(LeadResponse.model_validate(lead))

    return board


@router.post("", response_model=LeadResponse)
async def create_lead(
    data: LeadCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    if current_user.role == "call_agent":
        raise HTTPException(status_code=403, detail="call_agent cannot create leads")

    # Get or create contact
    if data.contact_id:
        contact = await db.get(Contact, data.contact_id)
        if not contact:
            raise HTTPException(status_code=404, detail="Contact not found")
    elif data.phone:
        # Check duplicate
        result = await db.execute(select(Contact).where(Contact.phone == data.phone))
        contact = result.scalar_one_or_none()
        if not contact:
            contact = Contact(name=data.name or "Unknown", phone=data.phone, email=data.email, source=data.source)
            db.add(contact)
            await db.flush()
    else:
        raise HTTPException(status_code=400, detail="Either contact_id or phone is required")

    assignee_id = data.assigned_to or current_user.id
    if current_user.role == "agent" and assignee_id != current_user.id:
        raise HTTPException(status_code=403, detail="Agent can only assign leads to self")

    lead = Lead(
        contact_id=contact.id,
        source=data.source,
        stage="new",
        lead_score=data.lead_score,
        budget_min=data.budget_min,
        budget_max=data.budget_max,
        property_type_interest=data.property_type_interest,
        location_preference=data.location_preference,
        timeline=data.timeline,
        assigned_to=assignee_id,
        priority=data.priority,
        stage_changed_at=datetime.utcnow(),
    )
    db.add(lead)
    await db.flush()

    await log_activity(db, lead.id, contact.id, "lead_created", f"Lead created manually by {current_user.name}", performed_by=current_user.id)

    if lead.assigned_to:
        await create_notification(
            db,
            lead.assigned_to,
            title=f"New lead assigned: {contact.name}",
            body=f"A new lead was created from {data.source} and assigned to you.",
            notif_type="new_lead",
            link=f"/leads/{lead.id}",
        )

    try:
        from app.services.followup_engine import schedule_followup_sequence
        await schedule_followup_sequence(
            db,
            lead.id,
            contact.id,
            trigger="new_lead",
            agent_id=lead.assigned_to,
        )
    except Exception:
        pass

    lead.priya_memory_brief = await build_memory_brief(db, lead, contact)

    await db.commit()
    await db.refresh(lead)

    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead.id)
    )
    return LeadResponse.model_validate(result.scalar_one())


@router.get("/{lead_id}", response_model=LeadResponse)
async def get_lead(lead_id: str, db: AsyncSession = Depends(get_db), current_user: Agent = Depends(get_current_user)):
    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)
    return LeadResponse.model_validate(lead)


@router.patch("/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: str, data: LeadUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    if current_user.role == "call_agent":
        update_fields = set(data.model_dump(exclude_unset=True).keys())
        forbidden_fields = {"assigned_to", "priority", "lead_score"}
        if update_fields & forbidden_fields:
            raise HTTPException(status_code=403, detail="call_agent cannot edit assignment or priority fields")

    contact = await db.get(Contact, lead.contact_id)

    for field, value in data.model_dump(exclude_unset=True, exclude={"personal_notes"}).items():
        setattr(lead, field, value)
    lead.updated_at = datetime.utcnow()

    # personal_notes goes on the contact
    if data.personal_notes:
        if contact:
            contact.personal_notes = data.personal_notes
            lead.priya_memory_brief = await build_memory_brief(db, lead, contact)

    if lead.assigned_to:
        await create_notification(
            db,
            lead.assigned_to,
            title="Lead info updated",
            body=f"{(contact.name if contact else 'Lead')} details were updated by {current_user.name}.",
            notif_type="reminder",
            link=f"/leads/{lead.id}",
        )

    await db.commit()
    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead_id)
    )
    return LeadResponse.model_validate(result.scalar_one())


@router.patch("/{lead_id}/stage", response_model=LeadResponse)
async def update_stage(
    lead_id: str, data: StageUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    if current_user.role == "call_agent":
        raise HTTPException(status_code=403, detail="call_agent cannot change lead stage")

    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)
    if data.stage == "lost" and not data.lost_reason:
        raise HTTPException(status_code=400, detail="lost_reason is required when marking a lead as lost")

    lead = await change_lead_stage(db, lead, data.stage, current_user.id, data.lost_reason)

    if lead.assigned_to:
        await create_notification(
            db,
            lead.assigned_to,
            title=f"Lead stage updated to {data.stage}",
            body=f"{current_user.name} moved this lead to {data.stage}.",
            notif_type="stage_change",
            link=f"/leads/{lead.id}",
        )

    await db.commit()

    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead_id)
    )
    return LeadResponse.model_validate(result.scalar_one())


@router.get("/{lead_id}/timeline", response_model=list[ActivityResponse])
async def get_timeline(lead_id: str, db: AsyncSession = Depends(get_db), current_user: Agent = Depends(get_current_user)):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    result = await db.execute(
        select(Activity)
        .options(selectinload(Activity.performed_by_agent))
        .where(Activity.lead_id == lead_id)
        .order_by(Activity.performed_at.desc())
    )
    return [ActivityResponse.model_validate(a) for a in result.scalars().all()]


@router.post("/{lead_id}/note", response_model=ActivityResponse)
async def add_note(
    lead_id: str, data: NoteCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    activity = await log_activity(
        db, lead_id, lead.contact_id,
        activity_type="note",
        title="Note added",
        description=data.description,
        performed_by=current_user.id,
    )
    contact = await db.get(Contact, lead.contact_id)
    lead.priya_memory_brief = await build_memory_brief(db, lead, contact)

    if lead.assigned_to:
        await create_notification(
            db,
            lead.assigned_to,
            title="New note added",
            body=f"{current_user.name} added a note on this lead.",
            notif_type="reminder",
            link=f"/leads/{lead.id}",
        )

    await db.commit()
    await db.refresh(activity)
    return ActivityResponse.model_validate(activity)


@router.post("/{lead_id}/call-log", response_model=ActivityResponse)
async def log_call(
    lead_id: str, data: CallLogCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    lead.last_contacted_at = datetime.utcnow()
    lead.call_count += 1
    lead.updated_at = datetime.utcnow()

    activity = await log_activity(
        db, lead_id, lead.contact_id,
        activity_type="call",
        title=f"Call logged — {data.outcome}",
        description=data.description,
        outcome=data.outcome,
        performed_by=current_user.id,
        meta={"duration_seconds": data.duration_seconds},
    )
    contact = await db.get(Contact, lead.contact_id)
    lead.priya_memory_brief = await build_memory_brief(db, lead, contact)

    if lead.assigned_to:
        await create_notification(
            db,
            lead.assigned_to,
            title="Call log updated",
            body=f"{current_user.name} logged a call outcome: {data.outcome}.",
            notif_type="reminder",
            link=f"/leads/{lead.id}",
        )

    await db.commit()
    await db.refresh(activity)
    return ActivityResponse.model_validate(activity)


@router.post("/{lead_id}/whatsapp")
async def send_whatsapp_message(
    lead_id: str, data: WhatsAppSend,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)
    contact = await db.get(Contact, lead.contact_id)

    result = await send_whatsapp(
        to_phone=contact.phone,
        template=data.template,
        variables={"name": contact.name, "agent_name": current_user.name, "custom_message": data.custom_message or ""},
        db=db,
        lead_id=lead_id,
        contact_id=lead.contact_id,
        agent_id=current_user.id,
    )
    await db.commit()
    return result


@router.post("/{lead_id}/notify")
async def notify_lead_contact(
    lead_id: str,
    data: LeadNotifyRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)
    contact = await db.get(Contact, lead.contact_id)
    if not contact:
        raise HTTPException(status_code=404, detail="Lead contact not found")

    allowed_channels = {"whatsapp", "email"}
    channels = [channel.lower() for channel in data.channels if channel]
    if not channels:
        raise HTTPException(status_code=400, detail="At least one channel is required")
    if any(channel not in allowed_channels for channel in channels):
        raise HTTPException(status_code=400, detail="Only whatsapp and email channels are supported")

    if data.scheduled_at:
        scheduled_utc = _normalize_to_utc_naive_from_local(data.scheduled_at)
        if scheduled_utc <= datetime.utcnow():
            raise HTTPException(status_code=400, detail="scheduled_at must be in the future")

        from app.models.followup import FollowUp

        for channel in channels:
            db.add(
                FollowUp(
                    lead_id=lead.id,
                    contact_id=contact.id,
                    agent_id=current_user.id,
                    channel=channel,
                    template="custom",
                    message_body=data.message,
                    subject=data.subject,
                    scheduled_at=scheduled_utc,
                    status="pending",
                    triggered_by="manual",
                )
            )

        await db.commit()
        return {
            "status": "scheduled",
            "channels": channels,
            "scheduled_at_utc": scheduled_utc.isoformat(),
            "count": len(channels),
        }

    results: dict[str, dict] = {}

    for channel in channels:
        if channel == "whatsapp":
            results[channel] = await send_whatsapp(
                to_phone=contact.phone,
                template="custom",
                variables={
                    "name": contact.name,
                    "agent_name": current_user.name,
                    "custom_message": data.message,
                },
                db=db,
                lead_id=lead.id,
                contact_id=contact.id,
                agent_id=current_user.id,
            )
            continue

        if channel == "email":
            if not contact.email:
                results[channel] = {"sent": False, "error": "Lead contact does not have an email"}
                continue

            results[channel] = await send_email(
                to_email=contact.email,
                template="custom",
                variables={
                    "subject": data.subject or f"Update from {current_user.name}",
                    "body": data.message,
                    "name": contact.name,
                    "agent_name": current_user.name,
                },
                db=db,
                lead_id=lead.id,
                contact_id=contact.id,
                agent_id=current_user.id,
            )

    await db.commit()
    return {
        "status": "sent",
        "results": results,
    }


@router.get("/{lead_id}/property-matches", response_model=list[PropertyResponse])
async def property_matches(
    lead_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    props = await find_matching_properties(
        db,
        budget_min=float(lead.budget_min) if lead.budget_min else None,
        budget_max=float(lead.budget_max) if lead.budget_max else None,
        property_type=lead.property_type_interest,
        location=lead.location_preference,
    )
    return [PropertyResponse.model_validate(p) for p in props]

@router.delete("/{lead_id}")
async def delete_lead(
    lead_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin or manager can delete leads")
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    from sqlalchemy import delete
    from app.models.models import Activity, Task, SiteVisit
    from app.models.followup import FollowUp
    await db.execute(delete(Activity).where(Activity.lead_id == lead_id))
    await db.execute(delete(Task).where(Task.lead_id == lead_id))
    await db.execute(delete(FollowUp).where(FollowUp.lead_id == lead_id))
    await db.execute(delete(SiteVisit).where(SiteVisit.lead_id == lead_id))
    await db.execute(delete(Lead).where(Lead.id == lead_id))
    await db.commit()
    return {"status": "deleted"}


# ─── MASTER PROFILE (Feature 3) ─────────────────────────────────────────────

@router.get("/{lead_id}/master-profile")
async def get_master_profile(
    lead_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Feature 3: Get full master profile for a lead."""
    result = await db.execute(
        select(Lead)
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .where(Lead.id == lead_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    contact = lead.contact
    profile = lead.master_profile or {}

    # Try to populate AI fields from campaign_leads if available
    from app.models.campaign_dashboard import CampaignLead as CDLead
    ai_data = {}
    if contact and contact.phone:
        phone_digits = contact.phone.replace("+", "").replace(" ", "")[-10:]
        try:
            phone_int = int(phone_digits) if phone_digits.isdigit() else None
        except (ValueError, TypeError):
            phone_int = None
        if phone_int:
            # Best-effort enrichment — never let a campaign-lookup error break the
            # whole master profile (e.g. for uploaded leads with no campaign data).
            try:
                cd_result = await db.execute(
                    select(CDLead)
                    .where(CDLead.phone_number == phone_int)
                    .order_by(CDLead.updated_at.desc())
                    .limit(1)
                )
                cd_lead = cd_result.scalar_one_or_none()
                if cd_lead:
                    ai_data = {
                        "config_preference": cd_lead.config_interest,
                        "budget_range": cd_lead.budget_signal,
                        "site_visit_intent": "Yes" if cd_lead.site_visit_committed else ("Maybe" if cd_lead.site_visit_timeframe else "No"),
                        "primary_language": cd_lead.language_preference,
                        "objection_type": cd_lead.objection_type,
                        "intent_level": cd_lead.intent_level,
                        "ai_summary": cd_lead.enriched_summary,
                        "key_quote": cd_lead.key_quote,
                    }
            except Exception:
                await db.rollback()
                ai_data = {}

    # Compute stats
    task_result = await db.execute(select(Task).where(Task.lead_id == lead_id))
    all_tasks = task_result.scalars().all()
    total_tasks = len(all_tasks)
    done_tasks = sum(1 for t in all_tasks if t.status == "done")
    completion_rate = (done_tasks / total_tasks * 100) if total_tasks > 0 else 0.0

    activities_result = await db.execute(
        select(Activity)
        .where(Activity.lead_id == lead_id)
        .order_by(Activity.performed_at.asc())
    )
    all_activities = activities_result.scalars().all()
    # Prefer explicit lead timestamps when available (more authoritative for recent syncs).
    first_contact = None
    if lead.created_at:
        first_contact = lead.created_at.isoformat()
    elif all_activities:
        first_contact = all_activities[0].performed_at.isoformat()

    # Use lead.last_contacted_at when present (set by call completion flows); fall back to activities.
    last_contact = None
    if getattr(lead, 'last_contacted_at', None):
        last_contact = lead.last_contacted_at.isoformat()
    elif all_activities:
        last_contact = all_activities[-1].performed_at.isoformat()

    days_in_pipeline = (datetime.utcnow() - lead.created_at).days if lead.created_at else 0

    return {
        # AI fields
        **ai_data,
        # Manual fields from stored profile
        "full_name": profile.get("full_name", contact.name if contact else None),
        "email": profile.get("email", contact.email if contact else None),
        "alternate_phone": profile.get("alternate_phone"),
        "city": profile.get("city"),
        "locality": profile.get("locality"),
        "occupation": profile.get("occupation"),
        "family_size": profile.get("family_size"),
        "current_living_situation": profile.get("current_living_situation"),
        "investment_purpose": profile.get("investment_purpose"),
        "source": profile.get("source", lead.source),
        "agent_notes": profile.get("agent_notes"),
        "priority_override": profile.get("priority_override"),
        "priority_override_reason": profile.get("priority_override_reason"),
        # Computed
        "total_calls": lead.call_count,
        "first_contact_date": first_contact,
        "last_contact_date": last_contact,
        "days_in_pipeline": days_in_pipeline,
        "completion_rate": round(completion_rate, 1),
        # Demographic fields — read from lead row columns first, fall back to master_profile JSON
        "age_range": lead.age_range or profile.get("age_range"),
        "occupation": lead.occupation or profile.get("occupation"),
        "occupation_other": lead.occupation_other,
        "family_size": lead.family_size or profile.get("family_size"),
        "income_range": lead.income_range,
        "property_budget": lead.property_budget,
        "preferred_location": lead.preferred_location,
        "purchase_timeline": lead.purchase_timeline,
        "last_call_status": lead.last_call_status,
        "last_call_interest": lead.last_call_interest,
        "last_call_topics": lead.last_call_topics,
    }


@router.patch("/{lead_id}/master-profile")
async def update_master_profile(
    lead_id: str,
    data: MasterProfileUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Feature 3: Update manual fields on the master profile."""
    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    profile = lead.master_profile or {}
    incoming = data.model_dump(exclude_unset=True)

    if not incoming:
        return {"status": "ok", "profile": profile}

    allowed_fields = {
        "config_preference", "budget_range", "site_visit_intent", "primary_language",
        "objection_type", "intent_level", "ai_summary", "key_quote",
        "full_name", "email", "alternate_phone", "city", "locality",
        "occupation", "family_size", "current_living_situation",
        "investment_purpose", "source", "agent_notes",
        "priority_override", "priority_override_reason",
    }

    if "priority_override" in incoming or "priority_override_reason" in incoming:
        if current_user.role not in {"admin", "manager"}:
            raise HTTPException(status_code=403, detail="Only admin/manager can override priority")

        next_override = incoming.get("priority_override", profile.get("priority_override"))
        next_reason = incoming.get("priority_override_reason", profile.get("priority_override_reason"))
        if next_override and (not isinstance(next_reason, str) or len(next_reason.strip()) < 20):
            raise HTTPException(status_code=400, detail="priority_override_reason must be at least 20 characters")

    for key, value in incoming.items():
        if key in allowed_fields:
            profile[key] = value

    lead.master_profile = profile
    lead.updated_at = datetime.utcnow()

    # Sync full_name and email back to contact
    contact = await db.get(Contact, lead.contact_id)
    if contact:
        if "full_name" in incoming and incoming["full_name"]:
            contact.name = incoming["full_name"]
        if "email" in incoming and incoming["email"]:
            contact.email = incoming["email"]

    await db.commit()
    return {"status": "ok", "profile": profile}


# ─── LEAD REASSIGNMENT (Feature 7) ─────────────────────────────────────────

class LeadReassignRequest(BaseModel):
    agent_id: str
    reason: str = Field(min_length=20)


@router.post("/{lead_id}/reassign")
async def reassign_lead(
    lead_id: str,
    data: LeadReassignRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Feature 7: Manually reassign a lead to another agent (admin/manager only)."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin/Manager only")

    lead = await db.get(Lead, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    assignee = await db.get(Agent, data.agent_id)
    if not assignee or not assignee.is_active:
        raise HTTPException(status_code=404, detail="Agent not found or inactive")

    old_assignee_id = lead.assigned_to
    old_assignee_name = None
    if old_assignee_id:
        old_assignee = await db.get(Agent, old_assignee_id)
        if old_assignee:
            old_assignee_name = old_assignee.name

    # Check if actually reassigning (different agent)
    if old_assignee_id == data.agent_id:
        return {"status": "ok", "message": "Lead already assigned to this agent"}

    lead.assigned_to = data.agent_id

    # Log activity
    contact = await db.get(Contact, lead.contact_id)
    activity = Activity(
        lead_id=lead.id,
        contact_id=lead.contact_id if contact else None,
        type="stage_change",
        title="Lead reassigned",
        description=f"Reassigned from {old_assignee_name or 'Unassigned'} to {assignee.name}. Reason: {data.reason}",
        performed_by=current_user.id,
        meta={
            "assignment_type": "manual_override",
            "previous_assignee": old_assignee_id,
            "new_assignee": data.agent_id,
            "reason": data.reason,
        }
    )
    db.add(activity)

    # Notify new assignee
    from app.services.lead_service import create_notification
    await create_notification(
        db,
        data.agent_id,
        title=f"Lead reassigned to you: {contact.name if contact else lead.id}",
        body=f"{current_user.name} reassigned this lead to you. Reason: {data.reason}",
        notif_type="new_lead",
        link=f"/leads/{lead.id}",
    )

    await db.commit()

    return {
        "status": "ok",
        "lead_id": lead.id,
        "previous_assignee": old_assignee_name,
        "new_assignee": assignee.name,
    }


# ─── GLOBAL ROUND-ROBIN LEAD DISTRIBUTION ───────────────────────────────────

class DistributeRequest(BaseModel):
    # Optional subset of agent ids to distribute to. If omitted, all active
    # call agents (and agents) are used.
    selected_agent_ids: Optional[list[str]] = None
    # If True (default) only assign leads that currently have no assignee.
    # If False, redistribute ALL leads evenly across the chosen agents.
    only_unassigned: bool = True
    # Optional explicit set of lead ids to distribute (e.g. a just-uploaded
    # batch). When provided, ONLY these leads are assigned — not the whole DB.
    lead_ids: Optional[list[str]] = None


@router.post("/distribute")
async def distribute_leads(
    payload: Optional[DistributeRequest] = None,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Round-robin (load-balanced) distribution of leads across call agents.

    Admin/manager only. Assigns each lead to the agent with the fewest active
    leads, so the workload stays balanced. By default only unassigned leads are
    distributed; set only_unassigned=false to rebalance everyone.
    """
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin/Manager only")

    from app.services.assignment_service import get_available_agents
    from app.services.lead_service import create_notification

    agents = await get_available_agents(db)
    if payload and payload.selected_agent_ids:
        wanted = set(payload.selected_agent_ids)
        agents = [a for a in agents if a.id in wanted]
    if not agents:
        raise HTTPException(status_code=400, detail="No active call agents to assign leads to")

    from sqlalchemy import update as sa_update

    only_unassigned = payload.only_unassigned if payload else True
    explicit_ids = payload.lead_ids if payload else None

    # Fetch only lead IDs (lightweight) — never load full ORM rows; there can be
    # thousands of leads and loading them all would time out.
    if explicit_ids:
        # Distribute ONLY this specific batch (e.g. a just-uploaded sheet).
        id_result = await db.execute(select(Lead.id).where(Lead.id.in_(explicit_ids)))
    elif only_unassigned:
        id_result = await db.execute(
            select(Lead.id).where(or_(Lead.assigned_to.is_(None), Lead.assigned_to == ""))
        )
    else:
        id_result = await db.execute(select(Lead.id))
    lead_ids = [row[0] for row in id_result.all()]

    if not lead_ids:
        return {"status": "ok", "assigned": 0, "message": "No leads to distribute"}

    # Seed each agent's current active-lead count so distribution stays balanced
    # against work they already hold.
    active_counts: dict[str, int] = {}
    for agent in agents:
        count_result = await db.execute(
            select(func.count(Lead.id)).where(
                Lead.assigned_to == agent.id,
                Lead.stage.notin_(["won", "lost", "nurture"]),
            )
        )
        active_counts[agent.id] = count_result.scalar() or 0

    # Compute the assignment in memory: each lead goes to the agent with the
    # fewest leads so far (load-balanced round-robin).
    buckets: dict[str, list[str]] = {a.id: [] for a in agents}
    for lid in lead_ids:
        assignee = min(agents, key=lambda a: active_counts[a.id])
        buckets[assignee.id].append(lid)
        active_counts[assignee.id] += 1

    # Apply with bulk UPDATEs in batches (fast — a few statements, not thousands).
    now = datetime.utcnow()
    BATCH = 500
    assigned_count = 0
    per_agent: dict[str, int] = {}
    for aid, ids in buckets.items():
        if not ids:
            continue
        for i in range(0, len(ids), BATCH):
            chunk = ids[i:i + BATCH]
            await db.execute(
                sa_update(Lead).where(Lead.id.in_(chunk)).values(assigned_to=aid, updated_at=now)
            )
        per_agent[aid] = len(ids)
        assigned_count += len(ids)

    await db.commit()

    # Create a follow-up call task per assigned lead so the call agent has
    # something to act on (and the call-completion form to fill after calling).
    all_assigned = [lid for ids in buckets.values() for lid in ids]
    if all_assigned:
        name_rows = await db.execute(
            select(Lead.id, Contact.name).join(Contact, Lead.contact_id == Contact.id)
            .where(Lead.id.in_(all_assigned))
        )
        names = {r[0]: r[1] for r in name_rows.all()}
        # Don't double up if the lead already has a pending task.
        existing_rows = await db.execute(
            select(Task.lead_id).where(Task.lead_id.in_(all_assigned), Task.status == "pending")
        )
        has_task = {r[0] for r in existing_rows.all()}
        for aid, ids in buckets.items():
            for lid in ids:
                if lid in has_task:
                    continue
                db.add(Task(
                    lead_id=lid,
                    title=f"Follow up: {names.get(lid) or 'lead'}",
                    task_type="call",
                    assigned_to=aid,
                    due_at=None,  # no due date → stays under Pending, not Overdue
                    priority="high",
                    status="pending",
                ))
        await db.commit()

    # Notify each agent who received leads.
    breakdown = []
    for agent in agents:
        count = per_agent.get(agent.id, 0)
        breakdown.append({"agent_id": agent.id, "agent_name": agent.name, "assigned": count})
        if count > 0:
            await create_notification(
                db,
                agent.id,
                title=f"{count} new lead(s) assigned to you",
                body=f"{current_user.name} distributed {count} lead(s) to you. Start calling!",
                notif_type="new_lead",
                link="/leads",
            )
    await db.commit()

    return {"status": "ok", "assigned": assigned_count, "agents": len(agents), "breakdown": breakdown}


class UnassignAgentRequest(BaseModel):
    agent_id: str


@router.post("/unassign-agent")
async def unassign_agent_leads(
    payload: UnassignAgentRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Remove all leads from a specific agent (admin/manager only).

    The leads are unassigned (returned to the pool, ready to re-distribute) and
    the agent's pending follow-up tasks for them are cancelled.
    """
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin/Manager only")

    from sqlalchemy import update as sa_update

    lead_rows = await db.execute(select(Lead.id).where(Lead.assigned_to == payload.agent_id))
    lead_ids = [r[0] for r in lead_rows.all()]
    if not lead_ids:
        return {"status": "ok", "unassigned": 0}

    now = datetime.utcnow()
    BATCH = 500
    for i in range(0, len(lead_ids), BATCH):
        chunk = lead_ids[i:i + BATCH]
        await db.execute(
            sa_update(Lead).where(Lead.id.in_(chunk)).values(assigned_to=None, updated_at=now)
        )
        await db.execute(
            sa_update(Task).where(
                Task.assigned_to == payload.agent_id,
                Task.lead_id.in_(chunk),
                Task.status == "pending",
            ).values(status="cancelled")
        )

    await db.commit()
    return {"status": "ok", "unassigned": len(lead_ids)}


# ─── SIMPLE LEAD UPLOAD (Name, Phone, Hot/Warm/Cold) ────────────────────────

def _pick_column(headers: list[str], *candidates: str) -> Optional[int]:
    """Return the index of the first header that matches any candidate keyword."""
    lowered = [(h or "").strip().lower() for h in headers]
    for i, h in enumerate(lowered):
        for c in candidates:
            if c in h:
                return i
    return None


def _map_score(raw: str) -> str:
    v = (raw or "").strip().lower()
    if "hot" in v:
        return "hot"
    if "cold" in v:
        return "cold"
    return "warm"


def _parse_lead_rows(content: bytes, filename: str) -> list[dict]:
    """Parse a CSV/XLSX with Name, Phone, and Type columns into lead dicts."""
    import csv as _csv
    import io as _io

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    table: list[list[str]] = []

    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl is required for Excel files")
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            table.append(["" if c is None else str(c) for c in row])
    elif ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        table = [list(r) for r in _csv.reader(_io.StringIO(text))]
    else:
        raise ValueError("Unsupported file type — upload a .xlsx or .csv file")

    # Drop fully-empty rows
    table = [r for r in table if any((c or "").strip() for c in r)]
    if not table:
        return []

    headers = table[0]
    name_i = _pick_column(headers, "name", "customer", "contact")
    phone_i = _pick_column(headers, "phone", "number", "mobile", "contact no")
    type_i = _pick_column(headers, "type", "category", "score", "label", "tier")

    # If headers don't look like headers (no phone column found), assume
    # positional layout: col0=name, col1=phone, col2=type, and include row 0.
    data_rows = table[1:]
    if phone_i is None:
        name_i, phone_i, type_i = 0, 1, 2
        data_rows = table

    leads: list[dict] = []
    for r in data_rows:
        def cell(idx):
            return (r[idx].strip() if idx is not None and idx < len(r) and r[idx] else "")
        phone_digits = "".join(ch for ch in cell(phone_i) if ch.isdigit())
        if len(phone_digits) < 10:
            continue
        phone = phone_digits[-10:]  # store the 10-digit number
        leads.append({
            "name": cell(name_i) or "Unknown",
            "phone": phone,
            "lead_score": _map_score(cell(type_i)),
        })
    return leads


@router.post("/upload")
async def upload_leads(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Upload leads from an Excel/CSV with columns: Name, Phone, Type (Hot/Warm/Cold).

    Admin/manager only. Creates unassigned leads — distribute them afterwards.
    """
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin/Manager only")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File is empty")

    try:
        rows = _parse_lead_rows(content, file.filename or "upload.xlsx")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found — need Name, Phone and Type columns")

    created = 0
    created_ids: list[str] = []
    for row in rows:
        # Reuse the contact if this phone already exists (phone is unique), else
        # create one. Either way we create a lead — the whole sheet is uploaded,
        # nothing is skipped as a "duplicate".
        existing = await db.execute(select(Contact).where(Contact.phone == row["phone"]))
        contact = existing.scalar_one_or_none()
        if not contact:
            contact = Contact(name=row["name"], phone=row["phone"], source="upload")
            db.add(contact)
            await db.flush()
        lead = Lead(
            contact_id=contact.id,
            source="manual",  # 'upload' isn't a valid lead_source enum value
            stage="new",
            lead_score=row["lead_score"],
            is_uploaded=True,  # marks this lead as part of an uploaded batch
            stage_changed_at=datetime.utcnow(),
        )
        db.add(lead)
        await db.flush()  # populate lead.id so we can return this batch
        created_ids.append(lead.id)
        created += 1

    await db.commit()
    # lead_ids lets the caller assign *only* this uploaded batch, not the whole DB.
    return {"status": "ok", "created": created, "skipped": len(rows) - created, "total": len(rows), "lead_ids": created_ids}


@router.post("/delete-uploaded")
async def delete_uploaded_leads(
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Delete every lead that was imported via the Staff page upload.

    Admin only. Upload-sourced leads are identified by their contact's
    source = 'upload'. Removes the leads, their contacts, and related records.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    from sqlalchemy import delete as sa_delete
    from app.models.models import Activity, Task, SiteVisit
    from app.models.followup import FollowUp

    # All leads from Staff uploads are flagged is_uploaded.
    lead_rows = await db.execute(select(Lead.id).where(Lead.is_uploaded == True))
    lead_ids = [r[0] for r in lead_rows.all()]

    BATCH = 500
    if lead_ids:
        for i in range(0, len(lead_ids), BATCH):
            chunk = lead_ids[i:i + BATCH]
            await db.execute(sa_delete(Activity).where(Activity.lead_id.in_(chunk)))
            await db.execute(sa_delete(Task).where(Task.lead_id.in_(chunk)))
            await db.execute(sa_delete(FollowUp).where(FollowUp.lead_id.in_(chunk)))
            await db.execute(sa_delete(SiteVisit).where(SiteVisit.lead_id.in_(chunk)))
            await db.execute(sa_delete(Lead).where(Lead.id.in_(chunk)))

    # Remove contacts created by uploads that no longer have any leads.
    contact_rows = await db.execute(select(Contact.id).where(Contact.source == "upload"))
    upload_contact_ids = [r[0] for r in contact_rows.all()]
    deleted_contacts = 0
    for cid in upload_contact_ids:
        remaining = await db.execute(select(func.count(Lead.id)).where(Lead.contact_id == cid))
        if (remaining.scalar() or 0) == 0:
            await db.execute(sa_delete(Contact).where(Contact.id == cid))
            deleted_contacts += 1

    await db.commit()
    return {"status": "ok", "deleted_leads": len(lead_ids), "deleted_contacts": deleted_contacts}


# ─── CAMPAIGN LEAD ASSIGNMENT TABLE ─────────────────────────────────────────

class CampaignAssignmentTableResponse(BaseModel):
    total: int
    page: int
    limit: int
    total_pages: int
    leads: list[dict]


class CampaignBulkAssignPayload(BaseModel):
    lead_ids: list[str] = Field(min_length=1)
    agent_id: str
    reason: Optional[str] = None


@router.get("/campaign/{campaign_id}/assignment-table")
async def campaign_assignment_table(
    campaign_id: str,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    priority_tier: Optional[str] = Query(None),
    assigned: Optional[str] = Query(None),
    agent_name: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query("asc"),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Get assignment table for campaign leads (admin/manager only)."""
    if current_user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only admin or manager can access campaign assignment table")

    filters = [Lead.campaign_id == campaign_id]

    if priority_tier:
        filters.append(Lead.priority == priority_tier)
    if assigned == "assigned":
        filters.append(Lead.assigned_to.isnot(None))
    elif assigned == "unassigned":
        filters.append(or_(Lead.assigned_to.is_(None), Lead.assigned_to == ""))

    if search:
        like = f"%{search.strip()}%"
        contact_result = await db.execute(select(Contact.id).where(
            or_(Contact.name.ilike(like), Contact.phone.ilike(like))
        ))
        contact_ids = [r[0] for r in contact_result.all()]
        if contact_ids:
            filters.append(Lead.contact_id.in_(contact_ids))
        else:
            filters.append(Lead.id == "none")

    total = await db.scalar(select(func.count(Lead.id)).where(*filters))

    order_col = Lead.created_at
    if sort_by == "priority":
        order_col = Lead.priority
    elif sort_by == "lead_score":
        order_col = Lead.lead_score

    direction = desc if sort_dir == "desc" else asc

    rows = (
        await db.scalars(
            select(Lead)
            .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
            .where(*filters)
            .order_by(direction(order_col))
            .offset((page - 1) * limit)
            .limit(limit)
        )
    ).all()

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit if total else 0,
        "leads": [
            {
                "id": lead.id,
                "name": lead.contact.name if lead.contact else "Unknown",
                "phone_number": lead.contact.phone if lead.contact else "N/A",
                "priority_tier": lead.priority,
                "lead_score": lead.lead_score,
                "assigned_agent": lead.assigned_agent.name if lead.assigned_agent else None,
                "intent_level": None,
                "dnd_flag": lead.dnd,
                "action_taken": lead.stage,
                "updated_at": lead.updated_at.isoformat() if lead.updated_at else None,
            }
            for lead in rows
        ],
    }


@router.post("/campaign/{campaign_id}/bulk-assign")
async def campaign_bulk_assign(
    campaign_id: str,
    payload: CampaignBulkAssignPayload,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Bulk-assign campaign leads to an agent (admin/manager only)."""
    if current_user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only admin or manager can assign campaign leads")

    agent = await db.get(Agent, payload.agent_id)
    if not agent or not agent.is_active:
        raise HTTPException(status_code=404, detail="Agent not found or inactive")

    result = await db.execute(
        select(Lead)
        .options(selectinload(Lead.contact), selectinload(Lead.assigned_agent))
        .where(
            Lead.campaign_id == campaign_id,
            Lead.id.in_(payload.lead_ids),
        )
    )
    leads = result.scalars().all()
    if not leads:
        raise HTTPException(status_code=404, detail="No matching leads found for this campaign")

    reassigned = 0
    updated = 0

    for lead in leads:
        if lead.assigned_to and lead.assigned_to != agent.id:
            reassigned += 1

    if reassigned > 0 and (not payload.reason or len(payload.reason) < 20):
        raise HTTPException(status_code=400, detail="Reassignment requires reason (min 20 chars)")

    for lead in leads:
        if lead.assigned_to == agent.id:
            continue

        previous_assignee = lead.assigned_to
        lead.assigned_to = agent.id
        lead.updated_at = datetime.utcnow()
        updated += 1

        db.add(
            Activity(
                lead_id=lead.id,
                type="assignment_update",
                description=f"Lead assigned to {agent.name} (bulk assignment)",
                performed_by=current_user.id,
                meta={
                    "assignment_type": "bulk_assign",
                    "previous_assignee": previous_assignee,
                    "new_assignee": agent.id,
                    "reason": payload.reason,
                },
            )
        )

    if updated > 0:
        await create_notification(
            db,
            agent.id,
            title=f"{updated} campaign leads assigned",
            body=f"{current_user.name} assigned {updated} lead(s) to you in campaign {campaign_id}." + (
                f" Reason: {payload.reason}" if payload.reason else ""
            ),
            notif_type="new_lead",
            link=f"/campaigns/{campaign_id}/dashboard",
        )

    await db.commit()

    return {
        "status": "ok",
        "assigned": updated,
        "reassigned": reassigned,
        "agent_name": agent.name,
        "agent_id": agent.id,
        "campaign_id": campaign_id,
    }


# ─── UPDATE LEAD PRIORITY (Admin only) ──────────────────────────────────────

class PriorityUpdateRequest(BaseModel):
    priority: str


@router.patch("/{lead_id}/priority")
async def update_lead_priority(
    lead_id: str,
    data: PriorityUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Update lead priority tier (admin/manager only)."""
    if current_user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Only admin or manager can update lead priority")

    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact)).where(Lead.id == lead_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    valid_priorities = ["P1", "P2", "P3", "P4", "P5", "high", "normal", "low"]
    if data.priority not in valid_priorities:
        raise HTTPException(status_code=400, detail=f"Invalid priority. Must be one of: {valid_priorities}")

    old_priority = lead.priority
    lead.priority = data.priority
    lead.updated_at = datetime.utcnow()

    # Log activity
    activity = Activity(
        lead_id=lead.id,
        contact_id=lead.contact_id,
        type="priority_change",
        title="Lead priority updated",
        description=f"Priority changed from {old_priority} to {data.priority} by {current_user.name}",
        performed_by=current_user.id,
        meta={"old_priority": old_priority, "new_priority": data.priority},
    )
    db.add(activity)

    # Notify assignee if exists
    if lead.assigned_to:
        from app.services.lead_service import create_notification
        await create_notification(
            db,
            lead.assigned_to,
            title=f"Lead priority updated to {data.priority}",
            body=f"{current_user.name} changed the priority of this lead from {old_priority} to {data.priority}.",
            notif_type="stage_change",
            link=f"/leads/{lead.id}",
        )

    await db.commit()

    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead_id)
    )
    return LeadResponse.model_validate(result.scalar_one())


# ─── DEMOGRAPHIC PROFILE ─────────────────────────────────────────────────────

@router.get("/{lead_id}/demographics", response_model=DemographicsResponse)
async def get_lead_demographics(
    lead_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Get demographic profile for a lead."""
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    return DemographicsResponse(
        age_range=lead.age_range,
        occupation=lead.occupation,
        occupation_other=lead.occupation_other,
        family_size=lead.family_size,
        income_range=lead.income_range,
        property_budget=lead.property_budget,
        preferred_location=lead.preferred_location,
        purchase_timeline=lead.purchase_timeline,
        last_call_status=lead.last_call_status,
        last_call_topics=lead.last_call_topics or [],
        last_call_interest=lead.last_call_interest,
    )


class DemographicsUpdateRequest(BaseModel):
    age_range: Optional[str] = None
    occupation: Optional[str] = None
    occupation_other: Optional[str] = None
    family_size: Optional[str] = None
    income_range: Optional[str] = None
    property_budget: Optional[str] = None
    preferred_location: Optional[str] = None
    purchase_timeline: Optional[str] = None
    last_call_status: Optional[str] = None
    last_call_interest: Optional[str] = None


@router.patch("/{lead_id}/demographics", response_model=LeadResponse)
async def update_lead_demographics(
    lead_id: str,
    data: DemographicsUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """
    Direct demographic edit. call_agent CANNOT use this endpoint —
    they must use the task completion form instead.
    """
    if current_user.role == "call_agent":
        raise HTTPException(status_code=403, detail="call_agent cannot directly edit demographics")

    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    _ensure_lead_scope(current_user, lead)

    for field, value in data.model_dump(exclude_unset=True).items():
        if value is not None:
            setattr(lead, field, value)

    lead.updated_at = datetime.utcnow()
    await db.commit()

    result = await db.execute(
        select(Lead).options(selectinload(Lead.contact), selectinload(Lead.assigned_agent)).where(Lead.id == lead_id)
    )
    return LeadResponse.model_validate(result.scalar_one())
