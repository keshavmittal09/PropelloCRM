"""
Bulk Task Ingest Router — CSV upload, ingestion, bulk assignment, and data export.

All endpoints are admin/manager only.
"""
import csv
import io
import re
import uuid
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.dependencies import get_current_user, get_db
from app.core.security import decode_token
from app.models.agent import Agent
from app.models.contact import Contact
from app.models.lead import Lead
from app.models.models import Activity, Task
from app.models.bulk_task_ingest import BulkTaskIngest, BulkTaskRecord
from app.schemas.schemas import (
    BulkTaskBulkAssignRequest,
    BulkTaskAssignByCallerRequest,
    BulkTaskIngestDetailResponse,
    BulkTaskIngestResponse,
    BulkTaskRecordResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter()


def _require_admin_or_manager(current_user: Agent) -> None:
    if current_user.role not in {"admin", "manager"}:
        raise HTTPException(status_code=403, detail="Admin/Manager only")


def _clean_phone(raw: str) -> str:
    """Normalize phone numbers to digits only, keeping country code."""
    if not raw:
        return ""
    digits = re.sub(r"[^\d]", "", str(raw))
    # If it's too long, strip the leading country code (91 for India)
    if len(digits) > 10 and digits.startswith("91"):
        digits = digits[2:]
    return digits


def _normalize_heat(heat: str) -> str:
    """Normalize lead heat bucket to hot/warm/cold."""
    if not heat:
        return "warm"
    heat = heat.strip().lower()
    if heat in ("hot", "h"):
        return "hot"
    elif heat in ("cold", "c"):
        return "cold"
    else:
        return "warm"


# ─── UPLOAD CSV ──────────────────────────────────────────────────────────────

@router.post("/upload", response_model=BulkTaskIngestResponse)
async def upload_bulk_csv(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """
    Upload a CSV file for bulk task ingestion.

    The CSV can have multiple sheets (tabs) representing different callers.
    For a flat CSV, the 'name' column in each row is used to group by caller.

    Supports both .csv and .xlsx files.
    """
    _require_admin_or_manager(current_user)

    filename = file.filename or "unknown.csv"
    content = await file.read()

    records_data = []
    caller_names_set = set()

    try:
        if filename.lower().endswith((".xlsx", ".xls")):
            # Handle Excel with multiple sheets
            try:
                import openpyxl
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="openpyxl is required for Excel files. Use CSV instead.",
                )
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                # First row is header
                headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
                for row in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = str(val) if val is not None else ""
                    row_dict["_caller_name"] = sheet_name.strip()
                    caller_names_set.add(sheet_name.strip())
                    records_data.append(row_dict)
            wb.close()
        else:
            # Handle CSV
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                # Normalize keys: lowercase + underscores
                cleaned = {
                    k.strip().lower().replace(" ", "_"): (v.strip() if v else "")
                    for k, v in row.items()
                    if k
                }
                # Use the first column that looks like a name/caller identifier
                # for tab-based bifurcation. The 'name' column usually contains lead name,
                # so we use it as the caller grouping.
                caller = cleaned.get("name", "").strip() or "Unknown"
                cleaned["_caller_name"] = caller
                caller_names_set.add(caller)
                records_data.append(cleaned)

    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Could not decode file. Ensure UTF-8 encoding.")
    except Exception as e:
        logger.error(f"CSV parse error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not records_data:
        raise HTTPException(status_code=400, detail="No data rows found in the uploaded file.")

    # Count heat buckets
    hot = warm = cold = 0
    for r in records_data:
        heat = _normalize_heat(
            r.get("lead_heat_bucket", "") or r.get("lead_heat", "") or r.get("heat", "")
        )
        if heat == "hot":
            hot += 1
        elif heat == "cold":
            cold += 1
        else:
            warm += 1

    # Create ingest batch
    ingest = BulkTaskIngest(
        batch_name=batch_name,
        file_name=filename,
        uploaded_by=current_user.id,
        total_records=len(records_data),
        hot_count=hot,
        warm_count=warm,
        cold_count=cold,
        caller_names=sorted(caller_names_set),
        status="processing",
    )
    db.add(ingest)
    await db.flush()

    # Create individual records
    created_leads = 0
    created_tasks = 0
    skipped = 0
    failed = 0

    # ─── Pre-load Contacts & Leads to avoid N+1 queries ───
    phones_in_batch = set()
    for row in records_data:
        p = _clean_phone(row.get("phone_number") or row.get("phone") or row.get("mobile") or row.get("mobile_number") or row.get("contact_number") or "")
        if p and len(p) >= 7:
            phones_in_batch.add(p)

    contact_by_phone = {}
    lead_by_contact_id = {}

    if phones_in_batch:
        existing_contacts_res = await db.execute(
            select(Contact).where(Contact.phone.in_(phones_in_batch))
        )
        for c in existing_contacts_res.scalars().all():
            contact_by_phone[c.phone] = c

        contact_ids = [c.id for c in contact_by_phone.values()]
        if contact_ids:
            existing_leads_res = await db.execute(
                select(Lead)
                .where(Lead.contact_id.in_(contact_ids))
                .where(Lead.stage.notin_(["won", "lost"]))
                .order_by(Lead.created_at.asc())
            )
            for l in existing_leads_res.scalars().all():
                lead_by_contact_id[l.contact_id] = l

    # Column name mapping (flexible to handle various CSV formats)
    COLUMN_MAP = {
        "call_id": ["call_id", "callid", "call_identifier"],
        "phone_number": ["phone_number", "phone", "phone_no", "mobile", "mobile_number", "contact_number"],
        "name": ["name", "lead_name", "customer_name", "contact_name"],
        "transcript_url": ["transcript_url", "transcript_u_r_l", "transcript", "transcript_link"],
        "recording_url": ["recording_url", "recording_u_r_l", "recording", "recording_link"],
        "extracted_entities": ["extracted_entities", "extracted_d_entities", "entities"],
        "call_eval_tags": ["call_eval_tags", "call_eval_tag", "eval_tag", "eval_tags", "call_evaluation_tags"],
        "summary": ["summary", "call_summary"],
        "call_conversation_quality": ["call_conversation_quality", "call_conv_ersation_quality", "conversation_quality", "quality"],
        "call_dialing_at": ["call_dialing_at", "call_diali_ng_at", "dialing_at", "dial_time"],
        "call_ringing_at": ["call_ringing_at", "call_ringi_ng_at", "ringing_at"],
        "user_picked_up": ["user_picked_up", "user_pic_ked_up", "picked_up"],
        "call_status": ["call_status", "call_stat_us", "status"],
        "duration": ["duration", "call_duration", "duration_bucket", "call_duration_seconds"],
        "lead_heat_bucket": ["lead_heat_bucket", "lead_hea_t", "lead_heat", "heat_bucket", "heat", "category"],
        "lead_heat_reason": ["lead_heat_reason", "lead_hea_t_reason", "heat_reason", "lead_heat_t_reason"],
    }

    def _get_field(row: dict, field_name: str) -> str:
        """Get a field value from a row, trying multiple possible column names."""
        for alias in COLUMN_MAP.get(field_name, [field_name]):
            if alias in row and row[alias]:
                return row[alias]
        return ""

    for row in records_data:
        try:
            phone_raw = _get_field(row, "phone_number")
            phone = _clean_phone(phone_raw)
            lead_name = _get_field(row, "name")
            caller_name = row.get("_caller_name", "Unknown")
            heat = _normalize_heat(_get_field(row, "lead_heat_bucket"))

            # Skip records with no phone number
            if not phone or len(phone) < 7:
                skipped += 1
                record = BulkTaskRecord(
                    ingest_id=ingest.id,
                    caller_name=caller_name,
                    name=lead_name,
                    phone_number=phone_raw,
                    call_id=_get_field(row, "call_id"),
                    lead_heat_bucket=heat,
                    ingestion_status="skipped",
                    extra_data={"skip_reason": "Invalid or missing phone number"},
                )
                db.add(record)
                continue

            # ─── Find or create Contact ───────────────────────────────
            if phone in contact_by_phone:
                contact = contact_by_phone[phone]
            else:
                contact = Contact(
                    id=str(uuid.uuid4()),
                    name=lead_name or "Unknown",
                    phone=phone,
                    type="buyer",
                    source="campaign",
                )
                db.add(contact)
                contact_by_phone[phone] = contact

            # ─── Find or create Lead ──────────────────────────────────
            if contact.id in lead_by_contact_id:
                lead = lead_by_contact_id[contact.id]
                # Update heat/score if provided
                lead.lead_score = heat
                lead.last_contacted_at = datetime.utcnow()
            else:
                lead = Lead(
                    id=str(uuid.uuid4()),
                    contact_id=contact.id,
                    source="campaign",
                    stage="new",
                    lead_score=heat,
                    priority="P2" if heat == "hot" else ("P3" if heat == "warm" else "P4"),
                )
                db.add(lead)
                lead_by_contact_id[contact.id] = lead
                created_leads += 1

            # ─── Create Task for follow-up ────────────────────────────
            task_title = f"Follow up: {lead_name or phone}"
            if heat == "hot":
                task_priority = "high"
                task_desc = f"🔥 HOT LEAD — {lead_heat_reason_text(row)}"
            elif heat == "cold":
                task_priority = "low"
                task_desc = f"❄️ Cold lead — {lead_heat_reason_text(row)}"
            else:
                task_priority = "normal"
                task_desc = f"Warm lead — {lead_heat_reason_text(row)}"

            task = Task(
                id=str(uuid.uuid4()),
                lead_id=lead.id,
                title=task_title,
                description=task_desc,
                task_type="call",
                priority=task_priority,
                status="pending",
                created_by=current_user.id,
            )
            db.add(task)
            created_tasks += 1

            # ─── Log Activity ─────────────────────────────────────────
            activity = Activity(
                lead_id=lead.id,
                contact_id=contact.id,
                type="campaign_call",
                title=f"Bulk CSV ingest: {_get_field(row, 'call_status') or 'imported'}",
                description=_get_field(row, "summary") or "Imported from CSV bulk upload",
                outcome=_get_field(row, "call_status"),
                recording_url=_get_field(row, "recording_url"),
                transcript=_get_field(row, "transcript_url"),
                call_summary=_get_field(row, "summary"),
                call_eval_tag=_get_field(row, "call_eval_tags"),
                performed_by=current_user.id,
                meta={
                    "source": "bulk_csv_ingest",
                    "batch_id": ingest.id,
                    "call_id": _get_field(row, "call_id"),
                    "duration": _get_field(row, "duration"),
                    "lead_heat": heat,
                    "lead_heat_reason": _get_field(row, "lead_heat_reason"),
                },
            )
            db.add(activity)

            # ─── Create BulkTaskRecord ────────────────────────────────
            record = BulkTaskRecord(
                ingest_id=ingest.id,
                caller_name=caller_name,
                name=lead_name,
                call_id=_get_field(row, "call_id"),
                phone_number=phone,
                transcript_url=_get_field(row, "transcript_url"),
                recording_url=_get_field(row, "recording_url"),
                extracted_entities=_get_field(row, "extracted_entities"),
                call_eval_tags=_get_field(row, "call_eval_tags"),
                summary=_get_field(row, "summary"),
                call_conversation_quality=_get_field(row, "call_conversation_quality"),
                call_dialing_at=_get_field(row, "call_dialing_at"),
                call_ringing_at=_get_field(row, "call_ringing_at"),
                user_picked_up=_get_field(row, "user_picked_up"),
                call_status=_get_field(row, "call_status"),
                duration=_get_field(row, "duration"),
                lead_heat_bucket=heat,
                lead_heat_reason=_get_field(row, "lead_heat_reason"),
                contact_id=contact.id,
                lead_id=lead.id,
                task_id=task.id,
                ingestion_status="ingested",
                extra_data={
                    k: v for k, v in row.items()
                    if k != "_caller_name" and k not in {
                        alias for aliases in COLUMN_MAP.values() for alias in aliases
                    }
                } or None,
            )
            db.add(record)

        except Exception as e:
            logger.error(f"Failed to process row: {e}")
            failed += 1
            record = BulkTaskRecord(
                ingest_id=ingest.id,
                caller_name=row.get("_caller_name", "Unknown"),
                name=_get_field(row, "name"),
                phone_number=_get_field(row, "phone_number"),
                call_id=_get_field(row, "call_id"),
                ingestion_status="failed",
                extra_data={"error": str(e)},
            )
            db.add(record)

    # Update ingest totals
    ingest.created_leads = created_leads
    ingest.created_tasks = created_tasks
    ingest.skipped_records = skipped
    ingest.failed_records = failed
    ingest.status = "completed"

    await db.commit()
    await db.refresh(ingest)

    return BulkTaskIngestResponse.model_validate(ingest)


def lead_heat_reason_text(row: dict) -> str:
    """Extract human-readable heat reason from row."""
    COLUMN_MAP_LOCAL = {
        "lead_heat_reason": ["lead_heat_reason", "lead_hea_t_reason", "heat_reason", "lead_heat_t_reason"],
    }
    for alias in COLUMN_MAP_LOCAL.get("lead_heat_reason", []):
        if alias in row and row[alias]:
            return row[alias]
    return "Imported from CSV"


# ─── LIST BATCHES ────────────────────────────────────────────────────────────

@router.get("/batches", response_model=list[BulkTaskIngestResponse])
async def list_batches(
    limit: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """List all bulk ingest batches."""
    _require_admin_or_manager(current_user)

    result = await db.execute(
        select(BulkTaskIngest)
        .order_by(BulkTaskIngest.created_at.desc())
        .limit(limit)
    )
    return [BulkTaskIngestResponse.model_validate(b) for b in result.scalars().all()]


# ─── GET BATCH DETAILS ───────────────────────────────────────────────────────

@router.get("/batch/{batch_id}", response_model=BulkTaskIngestDetailResponse)
async def get_batch_detail(
    batch_id: str,
    caller_name: Optional[str] = None,
    heat: Optional[str] = None,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Get batch details with filtered records."""
    _require_admin_or_manager(current_user)

    ingest = await db.get(BulkTaskIngest, batch_id)
    if not ingest:
        raise HTTPException(status_code=404, detail="Batch not found")

    query = select(BulkTaskRecord).where(BulkTaskRecord.ingest_id == batch_id)
    if caller_name:
        query = query.where(BulkTaskRecord.caller_name == caller_name)
    if heat:
        query = query.where(BulkTaskRecord.lead_heat_bucket == _normalize_heat(heat))

    query = query.order_by(BulkTaskRecord.created_at.desc())
    query = query.offset((page - 1) * limit).limit(limit)

    result = await db.execute(query)
    records = [BulkTaskRecordResponse.model_validate(r) for r in result.scalars().all()]

    return BulkTaskIngestDetailResponse(
        **BulkTaskIngestResponse.model_validate(ingest).model_dump(),
        records=records,
    )


# ─── DELETE BATCH ─────────────────────────────────────────────────────────────

@router.delete("/batch/{batch_id}")
async def delete_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Delete a batch and all its records. Does NOT delete the created leads/tasks."""
    _require_admin_or_manager(current_user)

    ingest = await db.get(BulkTaskIngest, batch_id)
    if not ingest:
        raise HTTPException(status_code=404, detail="Batch not found")

    await db.delete(ingest)
    await db.commit()
    return {"status": "deleted", "batch_id": batch_id, "batch_name": ingest.batch_name}


# ─── BULK ASSIGN TASKS ──────────────────────────────────────────────────────

@router.post("/batch/{batch_id}/bulk-assign")
async def bulk_assign_tasks(
    batch_id: str,
    data: BulkTaskBulkAssignRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Bulk assign selected records (and their tasks) to an agent."""
    _require_admin_or_manager(current_user)

    # Validate agent exists
    agent = await db.get(Agent, data.agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    result = await db.execute(
        select(BulkTaskRecord)
        .where(BulkTaskRecord.ingest_id == batch_id)
        .where(BulkTaskRecord.id.in_(data.record_ids))
    )
    records = result.scalars().all()

    assigned_count = 0
    for record in records:
        record.assigned_to = data.agent_id

        # Also assign the linked task
        if record.task_id:
            task = await db.get(Task, record.task_id)
            if task:
                task.assigned_to = data.agent_id

        # Also assign the linked lead
        if record.lead_id:
            lead = await db.get(Lead, record.lead_id)
            if lead:
                lead.assigned_to = data.agent_id

        assigned_count += 1

    await db.commit()
    return {
        "status": "ok",
        "assigned": assigned_count,
        "agent_id": data.agent_id,
        "agent_name": agent.name,
    }


@router.post("/batch/{batch_id}/assign-all")
async def assign_all_in_batch(
    batch_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Assign ALL records in a batch to a single agent (the whole ingest as one)."""
    _require_admin_or_manager(current_user)

    agent_id = data.get("agent_id")
    if not agent_id:
        raise HTTPException(status_code=400, detail="agent_id is required")

    agent = await db.get(Agent, agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    ingest = await db.get(BulkTaskIngest, batch_id)
    if not ingest:
        raise HTTPException(status_code=404, detail="Batch not found")

    result = await db.execute(
        select(BulkTaskRecord.task_id, BulkTaskRecord.lead_id)
        .where(BulkTaskRecord.ingest_id == batch_id)
    )
    rows = result.all()

    assigned_count = len(rows)
    task_ids = [r.task_id for r in rows if r.task_id]
    lead_ids = [r.lead_id for r in rows if r.lead_id]

    # Bulk update BulkTaskRecord
    await db.execute(
        update(BulkTaskRecord)
        .where(BulkTaskRecord.ingest_id == batch_id)
        .values(assigned_to=agent_id)
    )

    # Bulk update Tasks
    if task_ids:
        await db.execute(
            update(Task)
            .where(Task.id.in_(task_ids))
            .values(assigned_to=agent_id)
        )

    # Bulk update Leads
    if lead_ids:
        await db.execute(
            update(Lead)
            .where(Lead.id.in_(lead_ids))
            .values(assigned_to=agent_id)
        )

    await db.commit()
    return {
        "status": "ok",
        "assigned": assigned_count,
        "agent_id": agent_id,
        "agent_name": agent.name,
        "batch_name": ingest.batch_name,
    }


@router.post("/batch/{batch_id}/assign-by-caller")
async def assign_by_caller(
    batch_id: str,
    data: BulkTaskAssignByCallerRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Assign ALL records for a specific caller/tab name to an agent."""
    _require_admin_or_manager(current_user)

    agent = await db.get(Agent, data.agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    result = await db.execute(
        select(BulkTaskRecord)
        .where(BulkTaskRecord.ingest_id == batch_id)
        .where(BulkTaskRecord.caller_name == data.caller_name)
    )
    records = result.scalars().all()

    assigned_count = 0
    for record in records:
        record.assigned_to = data.agent_id

        if record.task_id:
            task = await db.get(Task, record.task_id)
            if task:
                task.assigned_to = data.agent_id

        if record.lead_id:
            lead = await db.get(Lead, record.lead_id)
            if lead:
                lead.assigned_to = data.agent_id

        assigned_count += 1

    await db.commit()
    return {
        "status": "ok",
        "assigned": assigned_count,
        "caller_name": data.caller_name,
        "agent_id": data.agent_id,
        "agent_name": agent.name,
    }


# ─── ADMIN DATA EXPORT ──────────────────────────────────────────────────────

@router.get("/export/{batch_id}")
async def export_batch_csv(
    batch_id: str,
    token: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Export batch data as CSV for admin download."""
    _require_admin_or_manager(current_user)

    ingest = await db.get(BulkTaskIngest, batch_id)
    if not ingest:
        raise HTTPException(status_code=404, detail="Batch not found")

    result = await db.execute(
        select(BulkTaskRecord)
        .where(BulkTaskRecord.ingest_id == batch_id)
        .order_by(BulkTaskRecord.caller_name, BulkTaskRecord.created_at)
    )
    records = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Caller/Tab", "Name", "Phone", "Call ID",
        "Lead Heat", "Heat Reason",
        "Call Status", "Duration", "Summary",
        "Call Eval Tags", "Call Quality",
        "Dialing At", "Ringing At", "User Picked Up",
        "Recording URL", "Transcript URL",
        "Extracted Entities",
        "Assigned To", "Lead ID", "Task ID", "Contact ID",
        "Ingestion Status", "Created At",
    ])

    for r in records:
        # Look up agent name if assigned
        agent_name = ""
        if r.assigned_to:
            agent_obj = await db.get(Agent, r.assigned_to)
            if agent_obj:
                agent_name = agent_obj.name

        writer.writerow([
            r.caller_name or "",
            r.name or "",
            r.phone_number or "",
            r.call_id or "",
            r.lead_heat_bucket or "",
            r.lead_heat_reason or "",
            r.call_status or "",
            r.duration or "",
            r.summary or "",
            r.call_eval_tags or "",
            r.call_conversation_quality or "",
            r.call_dialing_at or "",
            r.call_ringing_at or "",
            r.user_picked_up or "",
            r.recording_url or "",
            r.transcript_url or "",
            r.extracted_entities or "",
            agent_name,
            r.lead_id or "",
            r.task_id or "",
            r.contact_id or "",
            r.ingestion_status or "",
            str(r.created_at) if r.created_at else "",
        ])

    output.seek(0)
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", ingest.batch_name)
    filename = f"bulk_ingest_{safe_name}_{batch_id[:8]}.csv"

    return StreamingResponse(
        iter(["\ufeff", output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-all")
async def export_all_data(
    token: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: Agent = Depends(get_current_user),
):
    """Export ALL leads + tasks + activities as CSV for admin analysis."""
    _require_admin_or_manager(current_user)

    # Fetch leads with contacts
    leads_result = await db.execute(
        select(Lead)
        .options(
            selectinload(Lead.contact),
            selectinload(Lead.assigned_agent),
        )
        .order_by(Lead.created_at.desc())
    )
    leads = leads_result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Lead ID", "Contact Name", "Phone", "Email",
        "Source", "Stage", "Lead Score", "Priority",
        "Assigned Agent", "Campaign ID",
        "Budget Min", "Budget Max",
        "Property Interest", "Location Preference",
        "Call Count", "Last Contacted",
        "DND", "Last Remark",
        "Created At", "Updated At",
    ])

    for lead in leads:
        contact = lead.contact
        agent = lead.assigned_agent
        writer.writerow([
            lead.id,
            contact.name if contact else "",
            contact.phone if contact else "",
            contact.email if contact else "",
            lead.source,
            lead.stage,
            lead.lead_score,
            lead.priority,
            agent.name if agent else "",
            lead.campaign_id or "",
            str(lead.budget_min) if lead.budget_min else "",
            str(lead.budget_max) if lead.budget_max else "",
            lead.property_type_interest or "",
            lead.location_preference or "",
            lead.call_count,
            str(lead.last_contacted_at) if lead.last_contacted_at else "",
            "Yes" if lead.dnd else "No",
            lead.last_remark or "",
            str(lead.created_at),
            str(lead.updated_at),
        ])

    output.seek(0)
    now = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    return StreamingResponse(
        iter(["\ufeff", output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=propello_leads_export_{now}.csv"},
    )
